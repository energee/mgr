#!/usr/bin/env python3
"""Reconcile MGR orders from the Beer orders.xlsx source of truth.

MongoDB is deliberately not consulted. The script parses both historical
spreadsheet layouts, excludes internal/taproom rows, resolves approved
customer and brand aliases, prices every line from the workbook's Distributor
matrix, and performs an idempotent Supabase REST upsert.

Dry-run is the default. Pass ``--apply`` only after the plan has no unresolved
mappings. A JSON backup of the pre-image is written before any live mutation.
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
import unicodedata
import urllib.error
import urllib.parse
import urllib.request
import uuid
from collections import Counter, defaultdict
from dataclasses import asdict, dataclass
from datetime import date, datetime, timezone
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


IMPORT_NAMESPACE = uuid.UUID("3c39cd42-e9f2-45d0-b33d-d1429e79c7d2")
LEGACY_ORDER_NUMBER = re.compile(r"^ORD-\d{8}-[0-9a-f]{6}$")

CUSTOMER_ALIASES = {
    "sarene east": "Sarene Craft East",
    "sarene craft east": "Sarene Craft East",
    "sarane west": "Sarene Craft West",
    "sarene west": "Sarene Craft West",
    "sarene craft west": "Sarene Craft West",
    "sarene totals": "Sarene Craft Beer",
    "sarene totals pennsylvania": "Sarene Craft Beer",
    "sarene craft beer": "Sarene Craft Beer",
    "lake beverage": "Lake Beverage",
    "lake beverage rochester": "Lake Beverage",
    "hops express": "Hops Express - Hong Kong",
    "hops express hong kong": "Hops Express - Hong Kong",
    "coolship international": "Coolship International - South Korea",
    "coolship international south korea": "Coolship International - South Korea",
    "sixth city": "Sixth City - Ohio",
    "sixth city ohio": "Sixth City - Ohio",
    "china": "Odyssey Club - China",
    "odyssey club china": "Odyssey Club - China",
    "unusual holdings": "Unusual Holdings - Japan",
    "unusual holdings japan": "Unusual Holdings - Japan",
    "brew export": "Brew Export - UK",
    "brew export uk": "Brew Export - UK",
    "interbeer": "Interbeer - Europe",
    "interbeer europe": "Interbeer - Europe",
}

# User-approved aliases plus spelling/descriptor normalization found in the
# workbook. ``variant:`` selects a specific row when duplicate display names
# exist (Origen II has both a dormant duplicate and the active variant row).
BRAND_ALIASES = {
    "african wild dog animal 9": "African Wild Dog",
    "aurantia double": "Aurantia",
    "bluejay": "Blue Jay",
    "cardinal bird 2": "Cardinal",
    "ddh rubico": "Rubico II",
    "double lupula eggers": "Double Lupula (Eggers)",
    "eagle animal 8": "Eagle",
    "evergrain collab osprey": "Osprey",
    "great horned owl animal 11 tall trees collab": "Great Horned Owl",
    "hyena animal 6": "Hyena",
    "leopard ultra hopped ale": "Leopard",
    "lion animal 10": "Lion",
    "luminis braserie collab": "Luminis Brassarie Collab",
    "origen ii": "variant:origen-ii",
    "origen iii ultra hopped ale": "Origen III",
    "panther animal 7": "Panther",
    "parabellum offscript collab": "Parabellum",
    "priscus double ipa": "Priscus",
    "pumpkin stout": "Sever",
    "rubico iii ultra hopped ale": "Rubico III",
    "snow leopard animal 12": "Snow Leopard",
    "taupo": "Taupō",
    "taupo ii": "Taupō II",
    "taupo iii": "Taupō II",
    "tides iii": "Tides II",
    "vesparo": "Vespero",
    "verus ipa": "Verus",
    "vindobana": "Vindobona",
    "white tiger animal 13": "White Tiger",
    "zues": "Zeus",
}

# Historical sheets predate the explicit tier column. These defaults are
# derived from later workbook rows and the confirmed legacy product bands;
# an explicit workbook tier or a "Tier N" style label always wins.
DEFAULT_TIER_BY_BRAND = {
    "Akis": 3,
    "Akko": 3,
    "Aurantia": 5,
    "Capax": 5,
    "Costilla": 2,
    "Crest": 3,
    "Dōmbōc": 5,
    "Double Lupula": 5,
    "Double Prismatic": 5,
    "Edge": 4,
    "Edo": 1,
    "Glyphs": 3,
    "Gosev": 1,
    "Grandar": 2,
    "Hades": 6,
    "Hanno": 3,
    "Harvest": 5,
    "Hurron": 5,
    "Imara": 5,
    "Jantar": 2,
    "Kamo II": 5,
    "Lupula": 3,
    "Origen": 5,
    "Phokas": 3,
    "Priscus": 5,
    "Prismatic": 3,
    "Sever": 5,
    "Rein": 1,
    "Rooted": 1,
    "Rubico": 5,
    "Rubico II": 5,
    "Samo": 1,
    "Sonnen": 2,
    "Stark": 2,
    "Summer Lupula": 3,
    "Taupō": 3,
    "Taupō II": 3,
    "Tides II": 3,
    "Tides": 3,
    "Tribute": 4,
    "Tyrus": 5,
    "Verus": 3,
    "Vespero": 2,
    "Vindobona": 2,
    "Waves": 5,
    "Zeus": 5,
}


@dataclass
class RawLine:
    beer: str
    format_key: str
    quantity: int
    tier: int | None
    style_label: str


@dataclass
class RawOrder:
    sheet: str
    customer_label: str
    order_date: str
    requested_date: str | None
    lines: list[RawLine]


@dataclass
class PlannedLine:
    id: str
    brand_id: str
    brand_name: str
    selling_format_id: str
    format_key: str
    quantity: int
    tier: int
    unit_price: float
    keg_owner_id: str | None


@dataclass
class PlannedOrder:
    id: str
    order_number: str
    customer_id: str
    customer_name: str
    order_date: str
    requested_date: str | None
    status: str
    preserve_status: bool
    sheet: str
    lines: list[PlannedLine]


def normalized(value: object) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(char for char in text if not unicodedata.combining(char))
    text = text.lower().replace("&", " and ")
    return re.sub(r"[^a-z0-9]+", " ", text).strip()


def beer_base(value: object) -> str:
    return re.split(r"\s*\(", str(value or "").strip(), maxsplit=1)[0].strip()


def whole_quantity(value: object, *, sheet: str, cell: str) -> int:
    if value in (None, ""):
        return 0
    number = Decimal(str(value))
    if number != number.to_integral_value():
        raise ValueError(f"{sheet}!{cell}: quantity must be whole, got {value}")
    return int(number)


def iso_date(value: object) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return None


def parse_workbook(path: Path) -> tuple[list[RawOrder], dict[int, dict[str, Decimal]]]:
    workbook = load_workbook(path, data_only=True)
    price_sheet = workbook["Price Tiers"]
    prices: dict[int, dict[str, Decimal]] = {}
    for row in range(1, price_sheet.max_row + 1):
        tier_value = price_sheet.cell(row, 1).value
        if not isinstance(tier_value, (int, float)):
            continue
        tier = int(tier_value)
        if tier < 1 or tier > 7:
            continue
        prices[tier] = {
            "half": Decimal(str(price_sheet.cell(row, 3).value)),
            "sixtel": Decimal(str(price_sheet.cell(row, 4).value)),
            "case": Decimal(str(price_sheet.cell(row, 5).value)),
        }

    orders: list[RawOrder] = []
    for sheet in workbook.worksheets[1:]:
        first_row = [normalized(sheet.cell(1, column).value) for column in range(1, 11)]
        if len(first_row) > 1 and first_row[1] == "variant":
            legacy_rows: list[tuple[str, str, list[int]]] = []
            for row in range(2, sheet.max_row + 1):
                beer = sheet.cell(row, 2).value
                if not beer:
                    break
                order_date = iso_date(sheet.cell(row, 1).value)
                if not order_date:
                    raise ValueError(f"{sheet.title}!A{row}: missing order date")
                quantities = [
                    whole_quantity(
                        sheet.cell(row, column).value,
                        sheet=sheet.title,
                        cell=sheet.cell(row, column).coordinate,
                    )
                    for column in range(3, 9)
                ]
                legacy_rows.append((order_date, beer_base(beer), quantities))

            for customer, offset in (("Sarene East", 0), ("Sarene West", 3)):
                lines: list[RawLine] = []
                for _, beer, quantities in legacy_rows:
                    for format_key, quantity in zip(
                        ("half", "sixtel", "case"), quantities[offset : offset + 3]
                    ):
                        if quantity > 0:
                            lines.append(RawLine(beer, format_key, quantity, None, ""))
                if lines:
                    orders.append(
                        RawOrder(sheet.title, customer, legacy_rows[0][0], None, lines)
                    )
            continue

        headers: list[tuple[int, list[str]]] = []
        for row in range(1, min(sheet.max_row, 200) + 1):
            values = [normalized(sheet.cell(row, column).value) for column in range(1, 11)]
            if all(label in values for label in ("half", "sixtel", "case")):
                headers.append((row, values))

        for index, (header_row, values) in enumerate(headers):
            customer_label = str(sheet.cell(header_row, 3).value or "").strip()
            customer_key = normalized(customer_label)
            if (
                not customer_key
                or customer_key.startswith("customer totals")
                or "in house" in customer_key
                or "not active" in customer_key
                or customer_key == "blank"
            ):
                continue

            columns = {key: values.index(key) + 1 for key in ("half", "sixtel", "case")}
            tier_column = values.index("price tier") + 1 if "price tier" in values else None
            end_row = headers[index + 1][0] if index + 1 < len(headers) else header_row + 7
            lines: list[RawLine] = []
            order_dates: set[str] = set()
            requested_dates: set[str] = set()
            for row in range(header_row + 1, end_row):
                beer = sheet.cell(row, 3).value
                if not beer:
                    continue
                row_has_quantity = False
                tier_value = sheet.cell(row, tier_column).value if tier_column else None
                tier = int(tier_value) if isinstance(tier_value, (int, float)) else None
                style_label = str(sheet.cell(row, 4).value or "").strip()
                for format_key, column in columns.items():
                    quantity = whole_quantity(
                        sheet.cell(row, column).value,
                        sheet=sheet.title,
                        cell=sheet.cell(row, column).coordinate,
                    )
                    if quantity > 0:
                        row_has_quantity = True
                        lines.append(
                            RawLine(beer_base(beer), format_key, quantity, tier, style_label)
                        )
                if row_has_quantity:
                    if value := iso_date(sheet.cell(row, 1).value):
                        order_dates.add(value)
                    if value := iso_date(sheet.cell(row, 2).value):
                        requested_dates.add(value)

            if not lines:
                continue
            if len(order_dates) > 1 or len(requested_dates) > 1:
                raise ValueError(
                    f"{sheet.title} {customer_label}: mixed dates "
                    f"order={sorted(order_dates)} pickup={sorted(requested_dates)}"
                )
            sheet_date = datetime.strptime(sheet.title, "%m-%d-%y").date().isoformat()
            orders.append(
                RawOrder(
                    sheet.title,
                    customer_label,
                    next(iter(order_dates), sheet_date),
                    next(iter(requested_dates), None),
                    lines,
                )
            )

    return orders, prices


class SupabaseRest:
    def __init__(self, url: str, key: str):
        self.base = f"{url.rstrip('/')}/rest/v1"
        self.headers = {"apikey": key, "Authorization": f"Bearer {key}"}

    def request(
        self,
        method: str,
        table: str,
        *,
        query: str = "",
        body: object | None = None,
        prefer: str | None = None,
    ) -> Any:
        url = f"{self.base}/{table}{'?' + query if query else ''}"
        headers = dict(self.headers)
        payload = None
        if body is not None:
            payload = json.dumps(body, separators=(",", ":")).encode()
            headers["Content-Type"] = "application/json"
        if prefer:
            headers["Prefer"] = prefer
        request = urllib.request.Request(url, data=payload, headers=headers, method=method)
        try:
            with urllib.request.urlopen(request) as response:
                raw = response.read()
                return json.loads(raw) if raw else None
        except urllib.error.HTTPError as error:
            details = error.read().decode()
            raise RuntimeError(f"{method} {table}: HTTP {error.code}: {details}") from error

    def select(self, table: str, columns: str = "*") -> list[dict[str, Any]]:
        rows: list[dict[str, Any]] = []
        page_size = 1000
        while True:
            query = urllib.parse.urlencode(
                {"select": columns, "limit": str(page_size), "offset": str(len(rows))}
            )
            page = self.request("GET", table, query=query) or []
            rows.extend(page)
            if len(page) < page_size:
                return rows


def load_env(path: Path) -> None:
    for raw in path.read_text().splitlines():
        line = raw.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        os.environ.setdefault(key, value.strip().strip('"').strip("'"))


def resolve_tier(line: RawLine, canonical_brand: str, learned: dict[str, set[int]]) -> int:
    if line.tier is not None:
        return line.tier
    tier_match = re.search(r"tier\s*([1-7])", line.style_label, re.IGNORECASE)
    if tier_match:
        return int(tier_match.group(1))
    learned_tiers = learned.get(canonical_brand, set())
    if len(learned_tiers) == 1:
        return next(iter(learned_tiers))
    if canonical_brand in DEFAULT_TIER_BY_BRAND:
        return DEFAULT_TIER_BY_BRAND[canonical_brand]

    style = normalized(line.style_label)
    if style in {"pale", "pale ale", "ipa"}:
        return 3
    if style == "west coast":
        return 4
    if style in {"dipa", "double ipa", "tipa"}:
        return 5
    if style in {"mex lager", "pilsner", "vienna lager", "lager tier 2"}:
        return 2
    if style in {"lager", "helles"}:
        return 1
    raise ValueError(
        f"No price tier for {line.beer!r} ({canonical_brand}) style={line.style_label!r}"
    )


def chunks(values: list[Any], size: int = 100) -> list[list[Any]]:
    return [values[index : index + size] for index in range(0, len(values), size)]


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--apply", action="store_true")
    parser.add_argument("--env", type=Path, default=Path(".env"))
    parser.add_argument("--backup-dir", type=Path, default=Path("/tmp/mgr-beer-orders-backups"))
    args = parser.parse_args()

    load_env(args.env)
    url = os.environ.get("NEXT_PUBLIC_SUPABASE_URL")
    key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
    if not url or not key:
        raise RuntimeError("NEXT_PUBLIC_SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY are required")

    raw_orders, price_matrix = parse_workbook(args.workbook)
    if len(raw_orders) != 216 or sum(len(order.lines) for order in raw_orders) != 1205:
        raise ValueError(
            "Workbook shape changed: expected 216 external orders / 1205 nonzero lines, "
            f"got {len(raw_orders)} / {sum(len(order.lines) for order in raw_orders)}"
        )

    duplicate_keys = Counter(
        (order.order_date, normalized(order.customer_label)) for order in raw_orders
    )
    duplicates = [key for key, count in duplicate_keys.items() if count > 1]
    if duplicates:
        raise ValueError(f"Duplicate customer/date order blocks: {duplicates}")

    rest = SupabaseRest(url, key)
    brands = rest.select("brands", "id,name,variant")
    customers = rest.select("customers", "id,name,customer_type,sales_channel_id,price_tier_id")
    formats = rest.select("selling_formats", "id,name,unit_count,container:containers(name,type)")
    channels = rest.select("sales_channels", "id,name")
    keg_owners = rest.select("keg_owners", "id,name,is_active")
    existing_orders = rest.select(
        "orders",
        "id,order_number,customer_id,status,order_date,requested_date,scheduled_date,fulfilled_date,notes,is_export,created_at,updated_at",
    )
    existing_items = rest.select(
        "order_items",
        "id,order_id,brand_id,selling_format_id,quantity,unit_price,notes,created_at",
    )

    brands_by_name: dict[str, list[dict[str, Any]]] = defaultdict(list)
    brands_by_variant: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for brand in brands:
        brands_by_name[normalized(brand["name"])].append(brand)
        if brand.get("variant"):
            brands_by_variant[normalized(brand["variant"])].append(brand)

    def resolve_brand(source_name: str) -> dict[str, Any]:
        source_key = normalized(source_name)
        target = BRAND_ALIASES.get(source_key)
        if target and target.startswith("variant:"):
            matches = brands_by_variant[normalized(target.removeprefix("variant:"))]
        else:
            target_key = normalized(target or source_name)
            matches = brands_by_name[target_key]
            if not matches:
                matches = brands_by_variant[target_key]
        if len(matches) != 1:
            raise ValueError(
                f"Brand mapping {source_name!r} -> {target or source_name!r} resolved {len(matches)} rows"
            )
        return matches[0]

    customer_rows = {normalized(row["name"]): row for row in customers}
    distributor_channel = next(
        (row for row in channels if normalized(row["name"]) == "distributor"), None
    )
    if not distributor_channel:
        raise ValueError("Distributor sales channel not found")
    microstar_owner = next(
        (
            row
            for row in keg_owners
            if normalized(row["name"]) == "microstar" and row.get("is_active")
        ),
        None,
    )
    kegfleet_owner = next(
        (
            row
            for row in keg_owners
            if normalized(row["name"]) == "kegfleet" and row.get("is_active")
        ),
        None,
    )
    if not microstar_owner or not kegfleet_owner:
        raise ValueError("Active Microstar and KegFleet keg owners are required")

    customer_plan: dict[str, dict[str, Any]] = {}
    for raw in raw_orders:
        alias_key = re.sub(r"\s*\([^)]*\)\s*$", "", raw.customer_label).strip()
        canonical_name = CUSTOMER_ALIASES.get(normalized(alias_key))
        if not canonical_name:
            raise ValueError(f"No customer mapping for {raw.customer_label!r}")
        existing = customer_rows.get(normalized(canonical_name))
        customer_plan[canonical_name] = existing or {
            "id": str(uuid.uuid5(IMPORT_NAMESPACE, f"customer:{canonical_name}")),
            "name": canonical_name,
            "customer_type": "distributor",
            "sales_channel_id": distributor_channel["id"],
            "price_tier_id": None,
            "is_active": True,
            "notes": "Created from Beer orders.xlsx source-of-truth import",
        }

    format_plan: dict[str, dict[str, Any]] = {}
    for row in formats:
        container = row.get("container") or {}
        container_name = normalized(container.get("name"))
        if normalized(row.get("name")) == "case" and int(row.get("unit_count") or 0) == 24:
            format_plan["case"] = row
        elif container_name == "1 2 barrel":
            format_plan["half"] = row
        elif container_name == "1 6 barrel":
            format_plan["sixtel"] = row
    if set(format_plan) != {"half", "sixtel", "case"}:
        raise ValueError(f"Could not resolve all formats: {sorted(format_plan)}")

    resolved_raw: list[tuple[RawOrder, str, list[tuple[RawLine, dict[str, Any]]]]] = []
    learned_tiers: dict[str, set[int]] = defaultdict(set)
    for raw in raw_orders:
        canonical_customer = CUSTOMER_ALIASES[
            normalized(re.sub(r"\s*\([^)]*\)\s*$", "", raw.customer_label).strip())
        ]
        resolved_lines: list[tuple[RawLine, dict[str, Any]]] = []
        for line in raw.lines:
            brand = resolve_brand(line.beer)
            resolved_lines.append((line, brand))
            if line.tier is not None:
                learned_tiers[brand["name"]].add(line.tier)
        resolved_raw.append((raw, canonical_customer, resolved_lines))

    planned: list[PlannedOrder] = []
    today = date.today().isoformat()
    for raw, customer_name, resolved_lines in resolved_raw:
        customer = customer_plan[customer_name]
        key = (raw.order_date, normalized(customer_name))
        deterministic_id = str(uuid.uuid5(IMPORT_NAMESPACE, f"order:{key[0]}:{key[1]}"))
        slug = re.sub(r"[^A-Z0-9]+", "", normalized(customer_name).upper())[:12]
        order_number = f"XLSX-{raw.order_date.replace('-', '')}-{slug}"
        existing = next((row for row in existing_orders if row["order_number"] == order_number), None)
        order_id = str(existing["id"]) if existing else deterministic_id
        aggregate: dict[tuple[str, str, int, Decimal], int] = defaultdict(int)
        for line, brand in resolved_lines:
            tier = resolve_tier(line, brand["name"], learned_tiers)
            if tier not in price_matrix:
                raise ValueError(f"Tier {tier} is absent from Price Tiers sheet")
            unit_price = price_matrix[tier][line.format_key]
            aggregate[(brand["id"], brand["name"], tier, unit_price, line.format_key)] += line.quantity

        lines = []
        for (brand_id, brand_name, tier, unit_price, format_key), quantity in sorted(
            aggregate.items(), key=lambda entry: (entry[0][1], entry[0][4])
        ):
            format_id = str(format_plan[format_key]["id"])
            item_id = str(uuid.uuid5(IMPORT_NAMESPACE, f"item:{order_id}:{brand_id}:{format_id}"))
            keg_owner = (
                kegfleet_owner
                if "in house" in normalized(raw.customer_label)
                else microstar_owner
            )
            lines.append(
                PlannedLine(
                    item_id,
                    str(brand_id),
                    brand_name,
                    format_id,
                    format_key,
                    quantity,
                    tier,
                    float(unit_price),
                    str(keg_owner["id"])
                    if format_key in {"half", "sixtel"}
                    else None,
                )
            )

        target_status = "fulfilled" if raw.order_date < today else "scheduled"
        planned_order = PlannedOrder(
            order_id,
            order_number,
            str(customer["id"]),
            customer_name,
            raw.order_date,
            raw.requested_date,
            str(existing["status"]) if existing else target_status,
            bool(existing),
            raw.sheet,
            lines,
        )
        planned.append(planned_order)

    # Mongo-created rows are deliberately removed instead of matched or reused.
    # They lack workbook ownership data and some do not correspond to any
    # external spreadsheet order. The XLSX deterministic IDs are the only
    # identities retained by this importer.
    legacy = [
        row
        for row in existing_orders
        if LEGACY_ORDER_NUMBER.fullmatch(str(row["order_number"]))
        and not row.get("customer_id")
    ]

    print(
        json.dumps(
            {
                "mode": "APPLY" if args.apply else "DRY RUN",
                "workbook_orders": len(planned),
                "workbook_lines": sum(len(order.lines) for order in planned),
                "mongo_orders_to_delete": len(legacy),
                "existing_spreadsheet_orders": sum(order.preserve_status for order in planned),
                "new_orders": sum(not order.preserve_status for order in planned),
                "customers_to_create": sorted(
                    name for name, row in customer_plan.items() if normalized(name) not in customer_rows
                ),
                "status_counts": Counter(order.status for order in planned),
            },
            indent=2,
            default=lambda value: dict(value),
        )
    )

    if not args.apply:
        return 0

    args.backup_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    backup_path = args.backup_dir / f"beer-orders-preimage-{stamp}.json"
    backup_path.write_text(
        json.dumps(
            {
                "orders": existing_orders,
                "order_items": existing_items,
                "workbook": str(args.workbook),
                "created_at": datetime.now(timezone.utc).isoformat(),
            },
            indent=2,
        )
    )
    print(f"Backup: {backup_path}")

    new_customers = [
        row for name, row in customer_plan.items() if normalized(name) not in customer_rows
    ]
    if new_customers:
        rest.request(
            "POST",
            "customers",
            body=new_customers,
            prefer="return=representation",
        )

    legacy_ids = [str(row["id"]) for row in legacy]
    for id_batch in chunks(legacy_ids, 40):
        query = urllib.parse.urlencode({"order_id": f"in.({','.join(id_batch)})"})
        rest.request("DELETE", "order_items", query=query, prefer="return=minimal")
        query = urllib.parse.urlencode({"id": f"in.({','.join(id_batch)})"})
        rest.request("DELETE", "orders", query=query, prefer="return=minimal")

    existing_ids = {str(row["id"]) for row in existing_orders}
    new_order_rows = []
    for order in planned:
        if order.id in existing_ids:
            continue
        new_order_rows.append(
            {
                "id": order.id,
                "order_number": order.order_number,
                "customer_id": order.customer_id,
                "status": "draft",
                "order_date": order.order_date,
                "requested_date": order.requested_date,
                "notes": f"Beer orders.xlsx source of truth; sheet {order.sheet}",
                "is_export": order.customer_name
                not in {"Sarene Craft East", "Sarene Craft West", "Sarene Craft Beer", "Lake Beverage", "Sixth City - Ohio"},
            }
        )
    for batch in chunks(new_order_rows):
        rest.request("POST", "orders", body=batch, prefer="return=representation")

    for order in planned:
        query = urllib.parse.urlencode({"id": f"eq.{order.id}"})
        rest.request(
            "PATCH",
            "orders",
            query=query,
            body={
                "customer_id": order.customer_id,
                "order_date": order.order_date,
                "requested_date": order.requested_date,
                "scheduled_date": order.requested_date if order.status in {"scheduled", "picking", "packed"} else None,
                "notes": f"Beer orders.xlsx source of truth; sheet {order.sheet}",
                "is_export": order.customer_name
                not in {"Sarene Craft East", "Sarene Craft West", "Sarene Craft Beer", "Lake Beverage", "Sixth City - Ohio"},
            },
            prefer="return=minimal",
        )

    order_ids = [order.id for order in planned]
    for id_batch in chunks(order_ids, 40):
        query = urllib.parse.urlencode({"order_id": f"in.({','.join(id_batch)})"})
        rest.request("DELETE", "order_items", query=query, prefer="return=minimal")

    item_rows = []
    for order in planned:
        for line in order.lines:
            item_rows.append(
                {
                    "id": line.id,
                    "order_id": order.id,
                    "brand_id": line.brand_id,
                    "selling_format_id": line.selling_format_id,
                    "quantity": line.quantity,
                    "unit_price": line.unit_price,
                    "keg_owner_id": line.keg_owner_id,
                    "notes": f"Distributor price tier {line.tier}",
                }
            )
    for batch in chunks(item_rows):
        rest.request("POST", "order_items", body=batch, prefer="return=minimal")

    new_targets: dict[str, list[str]] = defaultdict(list)
    for order in planned:
        if order.id not in existing_ids:
            new_targets[order.status].append(order.id)
    transition_path = ["confirmed", "scheduled", "picking", "packed", "fulfilled"]
    target_rank = {state: index for index, state in enumerate(transition_path)}
    for state in transition_path:
        ids = [
            order_id
            for target, target_ids in new_targets.items()
            if target_rank.get(target, -1) >= target_rank[state]
            for order_id in target_ids
        ]
        for id_batch in chunks(ids, 40):
            query = urllib.parse.urlencode({"id": f"in.({','.join(id_batch)})"})
            body: dict[str, Any] = {"status": state}
            if state == "fulfilled":
                body["fulfilled_date"] = date.today().isoformat()
            rest.request("PATCH", "orders", query=query, body=body, prefer="return=minimal")

    final_orders = rest.select("orders", "id,customer_id,status,order_date")
    final_items = rest.select(
        "order_items", "id,order_id,brand_id,selling_format_id,keg_owner_id,quantity,unit_price"
    )
    planned_ids = {order.id for order in planned}
    imported_orders = [row for row in final_orders if str(row["id"]) in planned_ids]
    imported_items = [row for row in final_items if str(row["order_id"]) in planned_ids]
    checks = {
        "orders": len(imported_orders),
        "items": len(imported_items),
        "missing_customers": sum(not row.get("customer_id") for row in imported_orders),
        "missing_brands": sum(not row.get("brand_id") for row in imported_items),
        "missing_formats": sum(not row.get("selling_format_id") for row in imported_items),
        "zero_prices": sum(Decimal(str(row.get("unit_price") or 0)) <= 0 for row in imported_items),
        "missing_keg_owners": sum(
            not row.get("keg_owner_id")
            for row in imported_items
            if row.get("selling_format_id")
            in {format_plan["half"]["id"], format_plan["sixtel"]["id"]}
        ),
    }
    print("Validation:", json.dumps(checks, indent=2))
    expected_items = sum(len(order.lines) for order in planned)
    if checks != {
        "orders": 216,
        "items": expected_items,
        "missing_customers": 0,
        "missing_brands": 0,
        "missing_formats": 0,
        "zero_prices": 0,
        "missing_keg_owners": 0,
    }:
        raise RuntimeError(f"Post-import validation failed: {checks}")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as error:  # noqa: BLE001 - CLI boundary must fail loudly
        print(f"ERROR: {error}", file=sys.stderr)
        raise
